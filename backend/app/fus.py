from openpyxl import load_workbook



# ==========================================================
# NORMALIZAR
# ==========================================================

def normaliza(valor):

    if valor is None:

        return ""

    return str(valor).strip()



# ==========================================================
# LIMPIAR SERVICIO
# Línea: F30n-F30n -> F30n
# ==========================================================

def limpiar_servicio(valor):

    texto = normaliza(valor)


    if "-" in texto:

        texto = texto.split("-")[0]


    return texto.strip()





# ==========================================================
# LEER FUS
# Equivalente hoja Tipo de Día
# ==========================================================

def leer_fus(
    archivos_fus
):

    registros = []



    for archivo in archivos_fus:


        print("==============================")
        print("LEYENDO FUS:")
        print(archivo)
        print("==============================")


        wb = load_workbook(

            archivo,

            data_only=True

        )


        ws = wb.active



        if ws.max_column < 20:

            raise Exception(
                "El FUS no contiene las columnas A:T esperadas"
            )



        for fila in range(

            2,

            ws.max_row + 1

        ):



            # B = Tipo

            tipo = normaliza(

                ws.cell(
                    fila,
                    2
                ).value
                

            ).upper()



            # Aceptar los tres eventos
            if tipo not in ("EXP", "VPA", "VEX"):
                continue

            registro = {

                "tipo": tipo,

                # A Evento

                "evento":

                    ws.cell(
                        fila,
                        1
                    ).value,



                # C Inicio

                "hora":

                    ws.cell(
                        fila,
                        3
                    ).value,



                # D De

                "desde":

                    ws.cell(
                        fila,
                        4
                    ).value,



                # F A

                "hasta":

                    ws.cell(
                        fila,
                        6
                    ).value,



                # K Línea:
                # SERVICIO REAL

                "servicio":

                    limpiar_servicio(

                        ws.cell(
                            fila,
                            11
                        ).value

                    ),



                # I Bus

                "bus":

                    ws.cell(
                        fila,
                        9
                    ).value,



                # K Línea completa

                "linea":

                    ws.cell(
                        fila,
                        11
                    ).value,



                # L KM

                "km":

                    ws.cell(
                        fila,
                        12
                    ).value,



                # N Id

                "codigo":

                    ws.cell(
                        fila,
                        14
                    ).value,



                # R Tipo Mapeado

                "tipo_dia":

                    normaliza(

                        ws.cell(
                            fila,
                            18
                        ).value

                    ),



                # T Sentido

                "sentido":

                    ws.cell(
                        fila,
                        20
                    ).value,



                # Guardar fila completa A:T

                "fila":

                    [

                        ws.cell(
                            fila,
                            columna
                        ).value

                        for columna in range(
                            1,
                            21
                        )

                    ]

            }



            registros.append(
                registro
            )



        wb.close()



    print("==============================")
    print(
        "TOTAL EVENTOS:",
        len(registros)
    )
    print("==============================")


    return registros