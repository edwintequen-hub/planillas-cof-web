from pathlib import Path
from openpyxl import load_workbook

from .config import INFO_EXCEL



# ==========================================================
# NORMALIZAR UNIDAD
# ==========================================================

def normalizar_unidad(valor):

    if not valor:
        return None


    texto = str(valor).strip().upper()


    if "ALFA" in texto:

        return "Alfa"


    if "OMEGA" in texto:

        return "Omega"


    return str(valor).strip()



# ==========================================================
# NORMALIZAR TEXTO
# ==========================================================

def normaliza(valor):

    if valor is None:
        return ""

    return str(valor).strip().upper()





# ==========================================================
# CARGAR INFO
# ==========================================================

def cargar_info():


    datos = []


    wb = load_workbook(

        INFO_EXCEL,

        data_only=True

    )


    ws = wb.active



    encabezados = {}


    for col in range(1, ws.max_column + 1):

        valor = ws.cell(

            row=1,

            column=col

        ).value


        if valor:

            encabezados[

                normaliza(valor)

            ] = col



    col_unidad = encabezados.get(
        "UNIDAD"
    )

    col_servicio = encabezados.get(
        "SERVICIO"
    )

    col_terminal = encabezados.get(
        "TERMINAL"
    )

    col_tipo = encabezados.get(
        "TIPO DE DIA"
    )



    for fila in range(2, ws.max_row + 1):


        unidad_excel = ws.cell(

            fila,

            col_unidad

        ).value



        servicio = ws.cell(

            fila,

            col_servicio

        ).value



        terminal = ws.cell(

            fila,

            col_terminal

        ).value



        tipo_dia = None


        if col_tipo:

            tipo_dia = ws.cell(

                fila,

                col_tipo

            ).value



        if servicio:


            datos.append({

                "unidad_excel":

                    unidad_excel,


                "unidad":

                    normalizar_unidad(
                        unidad_excel
                    ),


                "servicio":

                    str(servicio).strip(),


                "terminal":

                    str(terminal).strip()
                    if terminal
                    else None,


                "tipo_dia":

                    tipo_dia

            })



    wb.close()


    return datos





# ==========================================================
# UNIDADES
# ==========================================================

def obtener_unidades():


    datos = cargar_info()


    unidades = sorted({

        x["unidad"]

        for x in datos

        if x["unidad"]

    })


    return unidades





# ==========================================================
# SERVICIOS POR UNIDAD
# ==========================================================

def obtener_servicios(unidad):


    datos = cargar_info()


    servicios = sorted({

        x["servicio"]

        for x in datos

        if x["unidad"] == unidad

    })


    return servicios





# ==========================================================
# TERMINAL POR UNIDAD Y SERVICIO
# ==========================================================

def obtener_terminal(

    unidad,

    servicio

):


    datos = cargar_info()


    servicio_buscar = normaliza(servicio)



    for x in datos:


        if (

            x["unidad"] == unidad

            and

            normaliza(
                x["servicio"]
            )
            == servicio_buscar

        ):

            return x["terminal"]



    return None





# ==========================================================
# NUEVO:
# UNIDAD DE UN SERVICIO
# ==========================================================

def obtener_unidad_por_servicio(servicio):


    datos = cargar_info()


    servicio_buscar = normaliza(servicio)



    unidades = set()



    for x in datos:


        if normaliza(

            x["servicio"]

        ) == servicio_buscar:


            unidades.add(

                x["unidad"]

            )



    if len(unidades) == 1:

        return unidades.pop()



    if len(unidades) > 1:

        return list(unidades)



    return None