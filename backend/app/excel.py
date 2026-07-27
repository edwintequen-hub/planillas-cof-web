from pathlib import Path
import shutil

from openpyxl import load_workbook


# ==========================================================
# COPIAR PLANTILLA
# ==========================================================

def copiar_plantilla(origen, destino):
    """
    Copia la plantilla Excel al archivo de salida.
    """

    origen = Path(origen)
    destino = Path(destino)

    destino.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    shutil.copy2(
        origen,
        destino,
    )


# ==========================================================
# ABRIR EXCEL
# ==========================================================

def abrir_excel(ruta):
    """
    Abre un libro Excel.
    """

    return load_workbook(
        filename=ruta,
    )


# ==========================================================
# GUARDAR EXCEL
# ==========================================================

def guardar_excel(workbook, ruta):
    """
    Guarda el libro.
    """

    workbook.save(ruta)
    workbook.close()


# ==========================================================
# OBTENER HOJA
# ==========================================================

def obtener_hoja(workbook, nombre):

    if nombre not in workbook.sheetnames:

        raise Exception(
            f"No existe la hoja '{nombre}'."
        )

    return workbook[nombre]


# ==========================================================
# ELIMINAR FILAS
# ==========================================================

def limpiar_desde_fila(
    hoja,
    fila_inicio,
    fila_fin,
):

    if fila_fin < fila_inicio:
        return

    for fila in range(
        fila_inicio,
        fila_fin + 1,
    ):

        for columna in range(
            1,
            hoja.max_column + 1,
        ):

            hoja.cell(
                row=fila,
                column=columna,
            ).value = None


# ==========================================================
# ULTIMA FILA CON DATOS
# ==========================================================

def ultima_fila(
    hoja,
    columna=1,
):

    fila = hoja.max_row

    while fila > 1:

        valor = hoja.cell(
            row=fila,
            column=columna,
        ).value

        if valor not in (
            "",
            None,
        ):
            return fila

        fila -= 1

    return 1


# ==========================================================
# COPIAR FILA
# ==========================================================

def copiar_fila(
    hoja_origen,
    fila_origen,
    hoja_destino,
    fila_destino,
    columna_inicio=1,
    columna_fin=None,
):

    if columna_fin is None:

        columna_fin = hoja_origen.max_column

    for columna in range(
        columna_inicio,
        columna_fin + 1,
    ):

        hoja_destino.cell(
            row=fila_destino,
            column=columna,
        ).value = hoja_origen.cell(
            row=fila_origen,
            column=columna,
        ).value