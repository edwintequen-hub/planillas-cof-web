from pathlib import Path
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .info import cargar_info
from .anexo4 import leer_anexo4

# ==========================================================
# NORMALIZAR TEXTO
# ==========================================================

def normaliza(valor):

    if valor is None:

        return ""

    return (
        str(valor)
        .strip()
        .upper()
    )



# ==========================================================
# SERVICIO PURO
#
# F30n -> F30N
# F30N-F30N -> F30N
# ==========================================================

def servicio_puro(valor):

    texto = normaliza(valor)


    if "-" in texto:

        texto = texto.split("-")[0]


    return texto.strip()



# ==========================================================
# FORMATO TIPO DIA
# ==========================================================

def formato_tipo_dia(valor):

    texto = normaliza(valor)


    if texto == "LABORAL":

        return "Laboral"



    if texto in (
        "SABADO",
        "SÁBADO"
    ):

        return "Sabado"



    if texto == "DOMINGO":

        return "Domingo"



    return str(valor)




# ==========================================================
# COLORES EVENTOS
# ==========================================================

COLOR_VPA = PatternFill(
    fill_type="solid",
    start_color="C6E0B4",
    end_color="C6E0B4"
)

COLOR_VEX = PatternFill(
    fill_type="solid",
    start_color="F4B183",
    end_color="F4B183"
)

# ==========================================================
# LEER FUS
#
# Columnas FUS:
#
# B  Tipo
# C  Inicio
# E  Fin
# J  Tipo Bus
# K  Servicio
# R  Tipo Mapeado
# T  Sentido
#
# Solo EXP
# ==========================================================

# ==========================================================
# NORMALIZAR HORA PARA ORDENAR
# ==========================================================
def hora_orden(valor):

    if valor is None:
        return ""

    if hasattr(valor, "strftime"):
        return valor.strftime("%H:%M:%S")

    return str(valor)


# ==========================================================
# LEER FUS
# ==========================================================
def leer_fus(
    archivos_fus
):


    registros = []



    for archivo in archivos_fus:



        print("==============================")
        print("LEYENDO FUS:")
        print(archivo)
        print("==============================")



        wb = load_workbook(
            archivo,
            data_only=True
        )


        ws = wb.active



        for fila in range(
            2,
            ws.max_row + 1
        ):



            tipo = normaliza(
                ws.cell(
                    fila,
                    2
                ).value
            )



            if tipo not in ("EXP", "VPA", "VEX"):

                continue




            registro = {

                "tipo": tipo,
                "evento":

                    ws.cell(
                        fila,
                        1
                    ).value,



                "hora":

                    ws.cell(
                        fila,
                        3
                    ).value,



                "fin":

                    ws.cell(
                        fila,
                        5
                    ).value,



                "tipo_bus":

                    ws.cell(
                        fila,
                        10
                    ).value,



                "servicio":

                    servicio_puro(

                        ws.cell(
                            fila,
                            11
                        ).value

                    ),



                "linea":

                    ws.cell(
                        fila,
                        11
                    ).value,



                "tipo_dia":

                    formato_tipo_dia(

                        ws.cell(
                            fila,
                            18
                        ).value

                    ),



                "sentido":

                    ws.cell(
                        fila,
                        20
                    ).value


            }



            registros.append(
                registro
            )



        wb.close()




    registros.sort(

    key=lambda x:(

        x["servicio"] or "",

        x["sentido"] or "",

        hora_orden(x["hora"])

        )

    )



    print("==============================")
    print(
        "TOTAL EXP:",
        len(registros)
    )
    print("==============================")


    return registros


# ==========================================================
# LIMPIAR BLOQUE
# Borra datos y formato de un rango de columnas
# ==========================================================

def limpiar_bloque(ws, fila_inicio, col_inicio, col_fin):

    ULTIMA_FILA_PLANTILLA = 555

    FILA_MODELO = 557

    for fila in range(fila_inicio, ULTIMA_FILA_PLANTILLA + 1):

        for col in range(col_inicio, col_fin + 1):

            celda = ws.cell(fila, col)

            modelo = ws.cell(FILA_MODELO, col)

            celda.value = None

            celda._style = copy(modelo._style)

        
# ==========================================================
# CREAR PLANILLA
#
# Mantiene estructura de Plantilla.xlsx
# ==========================================================

def crear_planilla(

    registros,

    plantilla,

    salida,

    unidad,

    servicio,

    tipo_dia,

    terminal

):
    import time

    inicio_planilla = time.time()

    wb = load_workbook(
        plantilla
    )


    ws = wb["Planilla"]




    # ======================================================
    # ENCABEZADOS
    # ======================================================

    ws["C3"] = terminal
    ws["AD3"] = terminal


    if str(servicio).isdigit():
        servicio_excel = int(servicio)
    else:
        servicio_excel = servicio

    ws["C4"] = servicio_excel
    ws["AD4"] = servicio_excel


    ws["C6"] = tipo_dia
    ws["AD6"] = tipo_dia




    fila_ida = 13

    fila_reg = 13




    # ======================================================
    # RECORRER REGISTROS
    # ======================================================

    print(f"Total registros recibidos: {len(registros)}")
    
    for registro in registros:




        if servicio_puro(
            registro["servicio"]
        ) != servicio_puro(servicio):

            continue





        if formato_tipo_dia(
            registro["tipo_dia"]
        ) != formato_tipo_dia(tipo_dia):

            continue





        sentido = str(
            registro["sentido"]
        )




        # ==================================================
        # IDA
        # ==================================================

        if sentido == "1":



            ws.cell(
                fila_ida,
                2
            ).value = registro["tipo_bus"]

            celda = ws.cell(fila_ida, 3)

            if registro["tipo"] == "VPA":
                celda.value = "Vacío a Patio IDA"
                celda.fill = COLOR_VPA

            elif registro["tipo"] == "VEX":
                celda.value = "Vacío a Cabezal IDA"
                celda.fill = COLOR_VEX

            else:
                celda.value = servicio + " IDA"



            ws.cell(
                fila_ida,
                7
            ).value = registro["hora"]



            ws.cell(
                fila_ida,
                21
            ).value = registro["fin"]




            fila_ida += 1






        # ==================================================
        # REGRESO
        # ==================================================

        elif sentido == "2":



            ws.cell(
                fila_reg,
                29
            ).value = registro["tipo_bus"]

            celda = ws.cell(fila_reg, 30)

            if registro["tipo"] == "VPA":

                celda.value = "Vacío a Patio REG"
                celda.fill = COLOR_VPA

            elif registro["tipo"] == "VEX":

                celda.value = "Vacío a Cabezal REG"
                celda.fill = COLOR_VEX

            else:

                celda.value = servicio + " REG"



            ws.cell(
                fila_reg,
                34
            ).value = registro["hora"]




            ws.cell(
                fila_reg,
                48
            ).value = registro["fin"]




            fila_reg += 1






    # ======================================================
    # VALIDAR DATOS
    # ======================================================

    if fila_ida == 13 and fila_reg == 13:


        wb.close()

        return None

    # ======================================================
    # LIMPIAR IDA
    # A:I
    # ======================================================

    limpiar_bloque(
        ws,
        fila_ida,
        1,
        9
    )

        # U:V
    limpiar_bloque(
        ws,
        fila_ida,
        21,
        23
    )

    # X:Y
    limpiar_bloque(
        ws,
        fila_ida,
        24,
        25
    )

    # ======================================================
    # LIMPIAR REG
    # ======================================================

    # AB:AJ
    limpiar_bloque(
        ws,
        fila_reg,
        28,
        36
    )

    # AV:AX
    limpiar_bloque(
        ws,
        fila_reg,
        48,
        50
    )

    # AY:AZ
    limpiar_bloque(
        ws,
        fila_reg,
        51,
        52
    )


    # ======================================================
    # GUARDAR
    # ======================================================

    salida.parent.mkdir(

        parents=True,

        exist_ok=True

    )




    wb.save(
        salida
    )



    wb.close()

    print(
        f"{servicio} {tipo_dia}: "
        f"{round(time.time() - inicio_planilla, 2)} segundos"
    )


    return salida
# ==========================================================
# GENERADOR GENERAL
#
# Maneja:
# - Servicio individual
# - Servicio TODO
# - Tipo Día individual
# - Tipo Día TODO
#
# Retorna resumen
# ==========================================================

def generar_planillas(

    archivos_fus,

    plantilla,

    info_excel,

    salida,

    unidad,

    servicio,

    tipo_dia,

    usar_anexo4=False

):


    # ======================================================
    # LEER ORIGEN DE DATOS
    # ======================================================

    if usar_anexo4:

        registros = leer_anexo4(
            archivos_fus[0]
        )

    else:

        registros = leer_fus(
            archivos_fus
        )


    info = cargar_info()


    # ======================================================
    # DICCIONARIO DE TERMINALES
    # ======================================================

    terminales = {}

    for dato in info:

        clave = (
            dato["unidad"],
            servicio_puro(dato["servicio"])
        )

        terminales[clave] = dato["terminal"]

    print("TERMINALES CARGADOS:")
    print(terminales)

    # ======================================================
    # OBTENER SERVICIOS
    # ======================================================

    if servicio == "TODO":


        servicios = sorted(

            list(

                set(

                    servicio_puro(
                        x["servicio"]
                    )

                    for x in registros

                    if x["servicio"]

                )

            )

        )


    else:


        servicios = [

            servicio_puro(
                servicio
            )

        ]






    # ======================================================
    # OBTENER TIPOS DIA
    # ======================================================

    if tipo_dia == "TODO":


        tipos_dia = [

            "Laboral",

            "Sabado",

            "Domingo"

        ]


    else:


        tipos_dia = [

            formato_tipo_dia(
                tipo_dia
            )

        ]






    contador = {


        "total":0,

        "laboral":0,

        "sabado":0,

        "domingo":0

    }

    # ======================================================
    # INDICE DE REGISTROS
    # ======================================================

    indice_registros = {}

    for registro in registros:

        clave = (
            servicio_puro(registro["servicio"]),
            formato_tipo_dia(registro["tipo_dia"])
        )

        if clave not in indice_registros:
            indice_registros[clave] = []

        indice_registros[clave].append(registro)

    print(f"Indice creado: {len(indice_registros)} combinaciones")



    # ======================================================
    # GENERAR
    # ======================================================

    for dia in tipos_dia:

        for serv in servicios:

            # ----------------------------------------------
            # BUSCAR TERMINAL
            # ----------------------------------------------

            terminal = terminales.get(
                (
                    unidad,
                    servicio_puro(serv)
                ),
                ""
            )

            if not terminal:
                print("SIN TERMINAL:", serv)

            # ----------------------------------------------
            # CARPETA DESTINO
            # ----------------------------------------------

            carpeta = (
                salida
                / unidad
                / dia
            )

            archivo_salida = carpeta / (
                serv + " " + dia + ".xlsx"
            )

            print(
                f"Generando: Servicio={serv} | Tipo={dia}"
            )

            resultado = crear_planilla(

                indice_registros.get(
                    (
                        servicio_puro(serv),
                        formato_tipo_dia(dia)
                    ),
                    []
                ),

                plantilla,
                archivo_salida,
                unidad,
                serv,
                dia,
                terminal

            )

            if resultado:

                print("GENERADA:", resultado)

                contador["total"] += 1

                if dia == "Laboral":
                    contador["laboral"] += 1

                elif dia == "Sabado":
                    contador["sabado"] += 1

                elif dia == "Domingo":
                    contador["domingo"] += 1

    return contador